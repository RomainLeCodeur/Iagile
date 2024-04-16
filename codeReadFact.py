import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import QUrl
from openpyxl import *
import factuality
import readability
from PyQt5.QtGui import QDesktopServices
import os
import fitz

class ChatInterface(QMainWindow):

    def open_pdf_guide(self):
        # Chemin relatif du fichier PDF
        file_path = "Guide_ReadFact.pdf"

        # Récupérer le chemin absolu du fichier PDF
        current_dir = os.path.dirname(__file__)
        file_path = os.path.join(current_dir, file_path)

        # Ouvrir le fichier PDF avec l'application par défaut
        QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))



# Thème Clair
BACKGROUND_COLOR = "#FFFFFF"
TEXT_COLOR = "#000000"
BUTTON_COLOR = "#4CAF50"
BUTTON_HOVER_COLOR = "#45a049"
BUTTON_TEXT_COLOR = "#FFFFFF"
CHECKBOX_COLOR = "#4CAF50"

API_URL = "https://api-inference.huggingface.co/models/valurank/en_readability"
headers = {"Authorization": "Bearer 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"}

class LanguageInterface:
    def __init__(self, chat_interface):
        self.chat_interface = chat_interface


class ChatInterface(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("ReadFact")
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.setStyleSheet(f"background-color: {BACKGROUND_COLOR}; color: {TEXT_COLOR};")

        layout = QGridLayout()
        
        # Zone de saisie des messages
        self.entry_field = QTextEdit()
        self.label_phrase = QLabel("Your sentence")
        layout.addWidget(self.label_phrase, 0, 0)
        layout.addWidget(self.entry_field, 1, 0)

        # Cases à cocher pour les fonctions
        self.function1_checkbox = QCheckBox("Factuality")
        self.function1_checkbox.setStyleSheet(f"QCheckBox {{ color: {TEXT_COLOR}; }}")  # Appliquer uniquement la couleur du texte
        self.function2_checkbox = QCheckBox("Readability")
        self.function2_checkbox.setStyleSheet(f"QCheckBox {{ color: {TEXT_COLOR}; }}")  # Appliquer uniquement la couleur du texte

        self.checkbox_layout = QVBoxLayout()
        self.checkbox_layout.addWidget(self.function1_checkbox)
        self.checkbox_layout.addWidget(self.function2_checkbox)
        self.checkbox_layout.setSpacing(0)  # Définir l'espacement entre les cases à cocher à 0

        layout.addLayout(self.checkbox_layout, 1, 1)

        # Console de sortie
        self.message_display = QTextEdit()
        self.message_display.setReadOnly(True)
        self.label_history = QLabel("History")
        layout.addWidget(self.label_history, 0, 2)
        layout.addWidget(self.message_display, 1, 2)

        # Boutons Send et Quit
        button_layout = QHBoxLayout()
        self.send_button = QPushButton("Send")
        self.send_button.clicked.connect(self.send_message)
        self.send_button.setStyleSheet(f"background-color: {BUTTON_COLOR}; color: {BUTTON_TEXT_COLOR};")
        button_layout.addWidget(self.send_button)

        self.quit_button = QPushButton("Quit")
        self.quit_button.clicked.connect(self.close)
        self.quit_button.setStyleSheet(f"background-color: {BUTTON_COLOR}; color: {BUTTON_TEXT_COLOR};")
        button_layout.addWidget(self.quit_button)
        
        layout.addLayout(button_layout, 2, 0)

        self.reset_button = QPushButton("Reset History")
        self.reset_button.clicked.connect(self.reset_history)
        self.reset_button.setStyleSheet(f"background-color: {BUTTON_COLOR}; color: {BUTTON_TEXT_COLOR};")
        button_layout.addWidget(self.reset_button)

        self.central_widget.setLayout(layout)

        # Création de la barre de menus
        self.create_menu()

        # Initialiser l'interface de langue
        self.language_interface = LanguageInterface(self)
    def reset_history(self):
        # Effacer l'historique dans la console de sortie
        self.message_display.clear()
        
    def create_menu(self):
        # Création de la barre de menus
        menubar = self.menuBar()

        # Menu Fichier
        file_menu = menubar.addMenu('File')

        # Action Enregistrer Historique
        save_history_action = QAction('Save History', self)
        save_history_action.triggered.connect(self.save_history)
        file_menu.addAction(save_history_action)
        
        # Action Charger Phrases depuis Excel
        load_phrases_action = QAction('Load From Excel', self)
        load_phrases_action.triggered.connect(self.load_phrases_from_excel)
        file_menu.addAction(load_phrases_action)

        # Action Charger Phrases depuis Excel
        load_phrases_action = QAction('Load From Pdf', self)
        load_phrases_action.triggered.connect(self.load_text_from_pdf)
        file_menu.addAction(load_phrases_action)

        # Menu Aide
        help_menu = menubar.addMenu('Help')

        # Action Aide PDF
        help_pdf_action = QAction('User Guide (PDF)', self)
        help_pdf_action.triggered.connect(self.open_pdf_guide)
        help_menu.addAction(help_pdf_action)

    def load_phrases_from_excel(self):
        # Demander à l'utilisateur de sélectionner le fichier Excel
        file_path, _ = QFileDialog.getOpenFileName(self, "Load From Excel", "", "Files Excel (*.xlsx)")

        if file_path:
            # Demander à l'utilisateur de spécifier les lignes et les colonnes
            rows_input, ok_rows = QInputDialog.getText(self, "Row", "Row numbers (separated by commas or range 'start;end'): ", QLineEdit.Normal, "")
            columns_input, ok_columns = QInputDialog.getText(self, "Columns", "Column letters (separated by commas or range 'start;end'): ", QLineEdit.Normal, "")
            
            if ok_rows and ok_columns:
                # Traitement des lignes et des colonnes spécifiées par l'utilisateur
                rows = self.process_rows_input(rows_input)
                columns = self.process_columns_input(columns_input)

                # Charger les phrases à partir du fichier Excel avec les lignes et colonnes spécifiées
                phrases = self.load_phrases(file_path, rows=rows, columns=columns)
                for phrase in phrases:
                    self.entry_field.append(phrase)

    def process_columns_input(self, columns_input):
        columns = []

        # Diviser les entrées par virgule
        inputs = columns_input.split(',')

        for item in inputs:
            if ';' in item:
                # Si l'entrée contient un point-virgule, traiter comme un intervalle
                start, end = item.split(';')
                columns.extend(self.get_column_range(start, end))
            else:
                # Sinon, traiter comme une lettre de colonne unique
                columns.append(item.strip())

        return columns

    def get_column_range(self, start, end):
        # Fonction pour générer une plage de lettres de colonne entre start et end inclusivement
        start_index = ord(start.upper()) - 65
        end_index = ord(end.upper()) - 65
        return [chr(i + 65) for i in range(start_index, end_index + 1)]

    def process_rows_input(self, rows_input):
        rows = []
        # Diviser les entrées par virgule
        inputs = rows_input.split(',')

        for item in inputs:
            if ';' in item:
                # Si l'entrée contient un point-virgule, traiter comme un intervalle
                start, end = item.split(';')
                rows.extend(range(int(start), int(end) + 1))
            else:
                # Sinon, traiter comme un numéro de ligne unique
                rows.append(int(item))
        return rows                

    def load_text_from_pdf(self):
        # Demander à l'utilisateur de sélectionner le fichier PDF
        file_path, _ = QFileDialog.getOpenFileName(self, "Load From PDF", "", "Files PDF (*.pdf)")

        if file_path:
            try:
                # Ouvrir le fichier PDF
                document = fitz.open(file_path)

                # Parcourir chaque page du PDF et extraire le texte
                text = ""
                for page in document:
                    text += page.get_text()

                # Ajouter le texte à la zone de saisie des messages
                self.entry_field.append(text)

                # Fermer le document PDF
                document.close()
            except Exception as e:
                # Gérer les erreurs lors de l'ouverture ou de l'extraction du texte du fichier PDF
                QMessageBox.critical(self, "Error", f"Error while loading PDF file : {e}")
    def load_phrases(self, file_path, rows=None, columns=None):
        phrases = []

        # Lire le fichier Excel et extraire les phrases
        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet = workbook.active

            if rows and columns:
                # Si des lignes et des colonnes sont spécifiées, ajoutez uniquement les cellules correspondantes
                for row in rows:
                    for column in columns:
                        cell_value = sheet[f"{column}{row}"].value
                        if cell_value:
                            phrases.append(str(cell_value))
            elif rows:
                # Si seulement des lignes sont spécifiées, ajoutez toutes les cellules de ces lignes
                for row in rows:
                    for cell in sheet[row]:
                        if cell.value:
                            phrases.append(str(cell.value))
            elif columns:
                # Si seulement des colonnes sont spécifiées, ajoutez toutes les cellules de ces colonnes
                for column in columns:
                    for cell in sheet[column]:
                        if cell.value:
                            phrases.append(str(cell.value))
            else:
                # Par défaut, ajoutez toutes les cellules non vides de la feuille de calcul
                for row in sheet.iter_rows(values_only=True):
                    for cell_value in row:
                        if cell_value:
                            phrases.append(str(cell_value))

            workbook.close()
        except Exception as e:
            # Gérer les erreurs de lecture du fichier Excel
            QMessageBox.critical(self, "Error", f"Error while loading Excel file : {e}")

        return phrases

    def save_history(self):
        # Demander à l'utilisateur de spécifier le nom et le chemin du fichier Excel
        file_path, _ = QFileDialog.getSaveFileName(self, "Save History", "", "Files Excel (*.xlsx)")

        if file_path:
            # Créer un nouveau classeur Excel
            wb = Workbook()
            ws = wb.active

            # Ajouter des en-têtes
            ws['A1'] = 'Sentence'
            ws['B1'] = 'Model'
            ws['C1'] = 'Score'
            ws['D1'] = 'Average Factuality'
            ws['E1'] = 'Average Readability'

            row_index = 2  # Commencer à la ligne suivante après l'en-tête
            messages = self.message_display.toPlainText().split('\n')

            # Initialiser des listes pour stocker les scores de chaque modèle
            factuality_scores = []
            readability_scores = []

            for message in messages:
                if message.startswith("Sentence:"):
                    phrase = message.split("Sentence:")[1].strip()
                elif message.startswith("Model:"):
                    model = message.split("Model:")[1].strip()
                elif message.startswith("Score:"):
                    score_str = message.split("Score:")[1].strip()

                    # Vérifier si le score contient un '/'
                    if '/' in score_str:
                        # Diviser le score en deux parties
                        factuality_score_str, readability_score_str = score_str.split('/')
                        
                        # Convertir les parties en float si elles ne sont pas vides
                        if factuality_score_str:
                            factuality_score_float = float(factuality_score_str.strip())
                            factuality_scores.append(factuality_score_float)

                        if readability_score_str:
                            readability_score_float = float(readability_score_str.strip())
                            readability_scores.append(readability_score_float)
                    else:
                        # Convertir le score en float si ce n'est pas vide
                        if score_str:
                            score_float = float(score_str)
                            factuality_scores.append(score_float) if model == 'Factuality' else readability_scores.append(score_float)
                        
                    ws.cell(row=row_index, column=1, value=phrase)
                    ws.cell(row=row_index, column=2, value=model)
                    ws.cell(row=row_index, column=3, value=score_str)

                    row_index += 1

            # Calculer et écrire la moyenne totale pour chaque modèle
            if factuality_scores:
                factuality_average = sum(factuality_scores) / len(factuality_scores)
                ws['D2'] = factuality_average

            if readability_scores:
                readability_average = sum(readability_scores) / len(readability_scores)
                ws['E2'] = readability_average

            # Enregistrer le classeur Excel
            wb.save(file_path)

            # Ouvrir le fichier Excel avec l'application par défaut
            QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))




    def parse_message(self, message):
        # Séparer le message en phrase, choix du modèle et score
        parts = message.split('\n')
        phrase = None
        model_choice = None
        score = None

        for part in parts:
            if part.startswith('Sentence:'):
                phrase = part.split('Sentence:')[-1].strip()
            elif part.startswith('Model:'):
                model_choice = part.split('Model:')[-1].strip()
            elif part.startswith('Score:'):
                score = part.split('Score:')[-1].strip()

        return phrase, model_choice, score
        
    def send_message(self):
        # Envoyer le message et afficher la réponse dans la console
        message = self.entry_field.toPlainText()

        if not self.function1_checkbox.isChecked() and not self.function2_checkbox.isChecked():
            QMessageBox.warning(self, "No Function Checked", "Please select a function.")
            return

        # Diviser le texte en phrases
        sentences = message.split('.')

        for sentence in sentences:
            if not sentence.strip():  # Ignorer les phrases vides
                continue
            sentence = sentence.replace('\n', '')

            self.display_message("Sentence: " + sentence)

            selected_function = None
            if self.function1_checkbox.isChecked() and self.function2_checkbox.isChecked():
                self.display_message("Model: Factuality and Readability")
                selected_function = 3
            elif self.function1_checkbox.isChecked() and not self.function2_checkbox.isChecked():
                self.display_message("Model: Factuality")
                selected_function = 1
            elif self.function2_checkbox.isChecked() and not self.function1_checkbox.isChecked():
                self.display_message("Model: Readability")
                selected_function = 2
#label1 factualité
#label0 opinion
            if selected_function == 1:
                response = factuality.scoreFact(sentence)
                if response['label']=="LABEL_1":
                    response=response['score']
                else:
                    response=1-response['score']
            elif selected_function == 2:
                response = readability.scoreRead(sentence)
            elif selected_function == 3:
                responseFact = factuality.scoreFact(sentence)
                if responseFact['label']=="LABEL_1":
                    responseFact=responseFact['score']
                else:
                    responseFact=1-responseFact['score']
                responseRead = readability.scoreRead(sentence)
                response = str(responseFact) + " / " + str(responseRead)

            self.display_message("Score: " + response)
        
        self.entry_field.clear()


    def display_message(self, message):
        # Afficher un message dans la console de sortie
        self.message_display.append(message)

    def open_pdf_guide(self):
        # Chemin relatif du fichier PDF
        file_path = "Guide_ReadFact.pdf"

        # Lancer l'application par défaut pour ouvrir le fichier PDF
        QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))

def main():
    # Lancer l'application
    app = QApplication(sys.argv)
    app.setStyle('Fusion')  # Utiliser le style Fusion pour une apparence plus moderne
    chat_interface = ChatInterface()
    chat_interface.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()